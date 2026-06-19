"""
excel_generator.py
Generates a comprehensive Labour Hire Price Calculator Excel file.
All sections mirror the html3.html Quote Tool form output exactly.
"""

import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

from brands import get_brand, brand_key

# ── Colour palette ─────────────────────────────────────────────
TEAL_DARK   = "1A6060"
TEAL_MID    = "2E8B8B"
TEAL_LIGHT  = "E6F7F7"
NAVY        = "17375E"
ORANGE      = "C0681A"
ORANGE_LITE = "FFF0D0"
YELLOW_HL   = "FFF8E1"
GREY_BG     = "F0F2F5"
GREY_LITE   = "F8F9FB"
WHITE       = "FFFFFF"
BLACK       = "1A1A1A"
DARK_GREY   = "4A5568"
GREEN_LITE  = "E8F8F0"
BLUE_LITE   = "E8E8FF"
RED_LITE    = "FFE8E8"
BORDER_CLR  = "C8D0DC"


# ── Style helpers ──────────────────────────────────────────────

def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)

def _font(bold=False, size=10, colour=BLACK, italic=False, name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=colour, italic=italic, name=name)

def _border(t="thin", colour=BORDER_CLR) -> Border:
    s = Side(style=t, color=colour)
    return Border(left=s, right=s, top=s, bottom=s)

def _thick_border() -> Border:
    th = Side(style="medium", color=TEAL_MID)
    tn = Side(style="thin",   color=BORDER_CLR)
    return Border(left=th, right=th, top=tn, bottom=tn)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _currency_fmt(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=float(value or 0))
    c.number_format = '$#,##0.00'
    c.font = _font(size=10)
    c.border = _border()
    c.alignment = _align("right")
    return c

def _num_fmt(ws, row, col, value, fmt="0.0"):
    c = ws.cell(row=row, column=col, value=float(value or 0))
    c.number_format = fmt
    c.font = _font(size=9)
    c.border = _border()
    c.alignment = _align("right")
    return c

def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


# ── Section / row builders ─────────────────────────────────────

def _section_banner(ws, row: int, text: str, col_span: int = 14):
    """Full-width orange section header."""
    c = ws.cell(row=row, column=1, value=text)
    c.font = _font(bold=True, colour=WHITE, size=11)
    c.fill = _fill(ORANGE)
    c.alignment = _align("left", "center")
    c.border = _border()
    _merge(ws, row, 1, row, col_span)
    ws.row_dimensions[row].height = 22


def _header_row(ws, row: int, labels: List[str], col_start: int = 1,
                bg: str = TEAL_DARK, fg: str = WHITE):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=lbl)
        c.fill  = _fill(bg)
        c.font  = _font(bold=True, colour=fg, size=9)
        c.alignment = _align("center")
        c.border = _border()
    ws.row_dimensions[row].height = 18


def _kv(ws, row: int, label: str, value: Any,
        lbl_col: int = 1, val_col: int = 3, val_end_col: int = 6,
        multiline: bool = False):
    """Key-value pair with label | value spanning columns."""
    lc = ws.cell(row=row, column=lbl_col, value=label)
    lc.font = _font(bold=True, size=9, colour=DARK_GREY)
    lc.fill = _fill(GREY_BG)
    lc.alignment = _align("left")
    lc.border = _border()
    _merge(ws, row, lbl_col, row, val_col - 1)

    vc = ws.cell(row=row, column=val_col, value=value)
    vc.font = _font(size=10)
    vc.border = _border()
    vc.alignment = _align("left", wrap=multiline)
    _merge(ws, row, val_col, row, val_end_col)

    if multiline and value:
        lines = str(value).count("\n") + 1
        ws.row_dimensions[row].height = max(18, 15 * lines)
    else:
        ws.row_dimensions[row].height = 18


# ── Cost calculation (mirrors JS calcRoleCost) ─────────────────

def _build_shift_map(shift_types: List[Dict]) -> Dict[str, Dict]:
    """shift_code → {hrs, is_night}"""
    m = {}
    for s in shift_types:
        desc = str(s.get("desc", "")).strip()
        if desc:
            m[desc] = {
                "hrs":      float(s.get("hrs", 0) or 0),
                "is_night": str(s.get("type", "DS")).upper() == "NS",
            }
    return m


def _num(role: Dict, *keys) -> float:
    """First non-empty numeric value among keys (legacy field fallback)."""
    for k in keys:
        v = role.get(k)
        if v not in (None, ""):
            try:
                return float(v)
            except (TypeError, ValueError):
                pass
    return 0.0


def _calc_role_cost(role: Dict, manning: Dict, shift_map: Dict,
                    dates: List[str], multiplier: float) -> Dict:
    """
    Mirror JS calcRoleCost() in shutdown mode: every hour is billed at the flat DS
    or NS rate. No weekday / weekend / public-holiday split and no ordinary /
    overtime / double-time tiers. Legacy field names (dsStd/dsRate/ds_rate) map
    onto the single flat DS/NS rate so old saved quotes still price.
    """
    qty = int(role.get("qty", 1) or 1)

    ds_rate = _num(role, "dsRate", "dsStd", "ds_rate")
    ns_rate = _num(role, "nsRate", "nsStd", "ns_rate")

    rid = str(role.get("id", ""))
    role_manning = manning.get(rid, {})

    ds_shifts = ns_shifts = 0
    ds_hrs = ns_hrs = 0.0

    for d in dates:
        code = role_manning.get(d, "OFF")
        if code in ("OFF", "—", ""):
            continue
        si = shift_map.get(code, {"hrs": 0, "is_night": False})
        hrs = si["hrs"] * qty
        if si["is_night"]:
            ns_shifts += 1
            ns_hrs += hrs
        else:
            ds_shifts += 1
            ds_hrs += hrs

    ds_cost = ds_hrs * ds_rate
    ns_cost = ns_hrs * ns_rate
    total   = ds_cost + ns_cost

    return {
        "ds_shifts":    ds_shifts,
        "ns_shifts":    ns_shifts,
        "ds_hrs":       ds_hrs,
        "ns_hrs":       ns_hrs,
        "ds_rate":      ds_rate,
        "ns_rate":      ns_rate,
        "ds_cost":      ds_cost,
        "ns_cost":      ns_cost,
        "total":        total,
        "client_charge": total * multiplier,
    }


# ── Main generator ─────────────────────────────────────────────

def _col_px(width_chars) -> int:
    """Approximate the pixel width of an Excel column from its character width."""
    return int(round((width_chars or 10) * 7 + 5))


def _add_brand_logo(ws, brand, col_w, target_h_px=56, row0=0):
    """Place the brand logo enlarged and horizontally centred across the used
    columns, anchored at the given 0-based row. Best-effort: silently skips if
    the logo can't be loaded so a quote still generates."""
    try:
        import io as _io
        import storage
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        from PIL import Image as PILImage
        data = storage.get_bytes(brand["logo_blob"])
        if not data:
            return
        iw, ih = PILImage.open(_io.BytesIO(data)).size
        h = target_h_px
        w = int(iw * (h / ih))
        px = [_col_px(col_w.get(c, 10)) for c in range(1, 28)]   # 27 used columns
        target_left = max(0, sum(px) / 2 - w / 2)                # centre the logo
        acc = 0
        anchor_col = 1
        for i, cw in enumerate(px, start=1):
            if acc + cw >= target_left:
                anchor_col = i
                break
            acc += cw
        img = XLImage(_io.BytesIO(data))
        img.width = w
        img.height = h
        img.anchor = f"{get_column_letter(anchor_col)}{row0 + 1}"
        ws.add_image(img)
    except Exception:
        pass


def generate_quote_excel(payload: Dict[str, Any]) -> bytes:
    """
    Generate a complete Labour Hire Quote Excel from the Quote Tool payload.
    Sections:
      A. Project Information
      B. Labour Rates & Requirements
      C. Manning Table
      D. Labour Allocation Details (cost breakdown per role)
      E. Pricing Summary
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Shutdown Quote"
    ws.sheet_view.showGridLines = False

    # ── Column widths ──────────────────────────────
    # Col 1=label, 2=spacer?, 3+=values / table data
    col_w = {1: 24, 2: 7, 3: 10}
    for c in range(4, 36):
        col_w[c] = 10
    for c in (6, 10, 11, 12):   # DS Cost / NS Cost / Total $ / Client $ columns
        col_w[c] = 13
    for col, w in col_w.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    di          = payload.get("data_input", {})
    shift_types = payload.get("shift_types", [])
    new_roles   = payload.get("new_roles",   [])
    staff_roles = payload.get("staff_roles", [])
    all_roles   = new_roles + staff_roles
    manning     = payload.get("manning",     {})
    dates       = payload.get("manning_dates", [])
    multiplier  = float(payload.get("multiplier", 1.35))
    oc          = payload.get("other_costs", {})
    summary     = payload.get("summary",    {})
    shift_map   = _build_shift_map(shift_types)

    row = 1

    # ══════════════════════════════════════════════
    # TITLE BLOCK
    # ══════════════════════════════════════════════
    brand     = get_brand(brand_key(payload))
    title_fg  = brand["xl_title_fg"]
    title_bg  = brand["xl_title_bg"]

    # Brand logo: enlarged and horizontally centred across the top of the sheet,
    # with the title and sub-line centred just below it.
    ws.row_dimensions[row].height = 46
    for _c in range(1, 28):
        ws.cell(row=row, column=_c).fill = _fill(title_bg)
    _add_brand_logo(ws, brand, col_w, target_h_px=56, row0=row - 1)
    row += 1

    ws.row_dimensions[row].height = 30
    c = ws.cell(row=row, column=1, value=brand["excel_title"])
    c.font      = Font(bold=True, size=16, color=title_fg, name="Calibri")
    c.alignment = _align("center", "center")
    c.fill      = _fill(title_bg)
    _merge(ws, row, 1, row, 27)
    row += 1

    ws.row_dimensions[row].height = 16
    generated = datetime.now().strftime("%d/%m/%Y %H:%M")
    job_id    = di.get("job_id", "")
    sub = ws.cell(row=row, column=1,
                  value=f"Job: {job_id}   |   Generated: {generated}   |   Multiplier: ×{multiplier:.2f}")
    sub.font      = _font(italic=True, colour=DARK_GREY, size=9)
    sub.fill      = _fill(title_bg)
    sub.alignment = _align("center", "center")
    _merge(ws, row, 1, row, 27)
    row += 2

    # ══════════════════════════════════════════════
    # SECTION A: PROJECT INFORMATION
    # ══════════════════════════════════════════════
    _section_banner(ws, row, "A.  PROJECT INFORMATION")
    row += 1

    kv_pairs_left = [
        ("Job #",               di.get("job_id", "")),
        ("Status",              di.get("status", "")),
        ("Client Job Number",   di.get("client_job_no", "")),
        ("Project / Job Title", di.get("proj_title", "")),
        ("Client Business Name",di.get("client_biz", "")),
        ("Site",                di.get("site", "")),
        ("Site Address",        di.get("site_addr", "")),
    ]
    kv_pairs_right = [
        ("Start Date",          di.get("start_date", "")),
        ("End Date",            di.get("end_date", "")),
        ("Quotation Validity",  di.get("validity", "")),
        ("Requesting Manager",  di.get("mgr_name", "")),
        ("Manager Role",        di.get("mgr_role", "")),
        ("Manager Contact",     di.get("mgr_phone", "")),
        ("Manager Email",       di.get("mgr_email", "")),
    ]

    max_rows = max(len(kv_pairs_left), len(kv_pairs_right))
    for i in range(max_rows):
        if i < len(kv_pairs_left):
            _kv(ws, row, kv_pairs_left[i][0],  kv_pairs_left[i][1],
                lbl_col=1, val_col=3, val_end_col=8)
        if i < len(kv_pairs_right):
            _kv(ws, row, kv_pairs_right[i][0], kv_pairs_right[i][1],
                lbl_col=9, val_col=11, val_end_col=17)
        row += 1

    # Client POC Email — full width
    client_email = di.get("client_email", "")
    if not client_email:
        # Try combined contact field format: "Name - email - phone"
        poc_raw = di.get("client_poc", di.get("client_contact", ""))
        client_email = poc_raw
    if client_email:
        _kv(ws, row, "Client POC / Email", client_email,
            lbl_col=1, val_col=3, val_end_col=17)
        row += 1

    # Scope of works — full width, multiline
    scope = di.get("scope", "")
    if scope:
        _kv(ws, row, "Scope of Works", scope,
            lbl_col=1, val_col=3, val_end_col=17, multiline=True)
        row += 1

    add_labour = di.get("add_labour", "")
    if add_labour:
        _kv(ws, row, "Additional Labour / Exclusions", add_labour,
            lbl_col=1, val_col=3, val_end_col=17, multiline=True)
        row += 1

    row += 1

    # ══════════════════════════════════════════════
    # SECTION B: LABOUR RATES & SHIFT TYPES
    # ══════════════════════════════════════════════
    _section_banner(ws, row, "B.  LABOUR RATES & SHIFT TYPES")
    row += 1

    # Multiplier
    lc = ws.cell(row=row, column=1, value="Client Multiplier")
    lc.font = _font(bold=True, size=9, colour=DARK_GREY)
    lc.fill = _fill(GREY_BG)
    lc.border = _border()
    _merge(ws, row, 1, row, 2)
    vc = ws.cell(row=row, column=3, value=multiplier)
    vc.number_format = '0.00"×"'
    vc.font = _font(bold=True, size=11, colour=ORANGE)
    vc.border = _border()
    vc.alignment = _align("center")
    ws.row_dimensions[row].height = 20
    row += 2

    # Shift Types table
    c = ws.cell(row=row, column=1, value="Shift Types")
    c.font = _font(bold=True, colour=ORANGE, size=10)
    _merge(ws, row, 1, row, 4)
    row += 1
    _header_row(ws, row, ["Code", "Hrs / Shift", "Type"], col_start=1)
    row += 1
    for s in shift_types:
        bg = BLUE_LITE if str(s.get("type", "DS")).upper() == "NS" else GREEN_LITE
        for col, val in enumerate([s.get("desc", ""), s.get("hrs", 0),
                                   s.get("type", "DS")], start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = _fill(bg)
            c.border = _border()
            c.alignment = _align("center" if col > 1 else "left")
            c.font = _font(size=9)
        row += 1
    row += 1

    # Role rate tables — shutdown mode: a single flat DS and NS $/hr per role.
    role_hdrs = ["Role", "Qty", "DS $/hr", "NS $/hr"]

    def _write_role_table(roles: List[Dict], title: str):
        nonlocal row
        c = ws.cell(row=row, column=1, value=title)
        c.font = _font(bold=True, colour=ORANGE, size=10)
        _merge(ws, row, 1, row, len(role_hdrs))
        row += 1
        # header row
        for i, lbl in enumerate(role_hdrs):
            hc = ws.cell(row=row, column=1 + i, value=lbl)
            hc.fill = _fill(TEAL_DARK)
            hc.font = _font(bold=True, colour=WHITE, size=9)
            hc.alignment = _align("center", wrap=True)
            hc.border = _border()
        ws.row_dimensions[row].height = 26
        row += 1
        for r in roles:
            vals = [
                r.get("role", ""), r.get("qty", 1),
                _num(r, "dsRate", "dsStd", "ds_rate"),
                _num(r, "nsRate", "nsStd", "ns_rate"),
            ]
            for i, val in enumerate(vals):
                cell = ws.cell(row=row, column=1 + i, value=val)
                cell.border = _border()
                if i == 0:
                    cell.font = _font(size=10)
                    cell.alignment = _align("left")
                elif i == 1:
                    cell.number_format = "0"
                    cell.font = _font(size=10)
                    cell.alignment = _align("center")
                else:
                    cell.number_format = '$#,##0.00'
                    cell.font = _font(size=10)
                    cell.alignment = _align("right")
                    if float(val or 0) == 0:
                        cell.font = _font(size=10, colour="AAAAAA")
            row += 1
        row += 1

    if new_roles:
        _write_role_table(new_roles, "New Roles / Ad-Hoc Positions")
    if staff_roles:
        _write_role_table(staff_roles, "Staffing Requirements")

    # ══════════════════════════════════════════════
    # SECTION C: MANNING TABLE
    # ══════════════════════════════════════════════
    if dates and all_roles:
        _section_banner(ws, row, "C.  MANNING TABLE", col_span=27)
        row += 1

        # Set narrow width for date columns
        for i in range(len(dates)):
            ws.column_dimensions[get_column_letter(2 + i)].width = 8
        # Client $ column
        client_col = 2 + len(dates)
        ws.column_dimensions[get_column_letter(client_col)].width = 13

        # Header row
        ws.row_dimensions[row].height = 32
        rh = ws.cell(row=row, column=1, value="Role")
        rh.fill = _fill(TEAL_DARK)
        rh.font = _font(bold=True, colour=WHITE, size=9)
        rh.border = _border()
        rh.alignment = _align("center")

        for i, d in enumerate(dates):
            try:
                dt = datetime.strptime(d, "%Y-%m-%d")
                lbl = dt.strftime("%a\n%d/%m")
            except Exception:
                lbl = d
            ch = ws.cell(row=row, column=2 + i, value=lbl)
            ch.fill = _fill(TEAL_MID)
            ch.font = _font(bold=True, colour=WHITE, size=8)
            ch.alignment = _align("center", wrap=True)
            ch.border = _border()

        tc = ws.cell(row=row, column=client_col, value="Client $")
        tc.fill = _fill(NAVY)
        tc.font = _font(bold=True, colour=WHITE, size=9)
        tc.border = _border()
        tc.alignment = _align("right")
        row += 1

        # Manning rows
        for r in all_roles:
            rid = str(r.get("id", ""))
            cost = _calc_role_cost(r, manning, shift_map, dates, multiplier)

            rname = ws.cell(row=row, column=1,
                            value=f"{r.get('role', '')} ×{r.get('qty', 1)}")
            rname.fill = _fill(TEAL_LIGHT)
            rname.font = _font(bold=True, colour=TEAL_DARK, size=9)
            rname.border = _border()

            for i, d in enumerate(dates):
                shift = manning.get(rid, {}).get(d, "—")
                sc = ws.cell(row=row, column=2 + i, value=shift)
                sc.border = _border()
                sc.alignment = _align("center")
                sc.font = _font(size=8)
                if shift.startswith("NS"):
                    sc.fill = _fill(BLUE_LITE)
                elif shift not in ("OFF", "—", ""):
                    sc.fill = _fill(GREEN_LITE)
                else:
                    sc.fill = _fill(GREY_LITE)

            cc = ws.cell(row=row, column=client_col,
                         value=cost["client_charge"])
            cc.number_format = '$#,##0.00'
            cc.font = _font(bold=True, size=9, colour=NAVY)
            cc.border = _border()
            cc.alignment = _align("right")
            row += 1

        # Total row
        total_client = sum(
            _calc_role_cost(r, manning, shift_map, dates, multiplier)["client_charge"]
            for r in all_roles
        )
        tr = ws.cell(row=row, column=1, value="TOTAL CLIENT CHARGE")
        tr.fill = _fill(NAVY)
        tr.font = _font(bold=True, colour=WHITE, size=10)
        tr.border = _border()
        _merge(ws, row, 1, row, client_col - 1)
        tc2 = ws.cell(row=row, column=client_col, value=total_client)
        tc2.number_format = '$#,##0.00'
        tc2.font = _font(bold=True, colour="FFE699", size=11)
        tc2.fill = _fill(NAVY)
        tc2.border = _border()
        tc2.alignment = _align("right")
        ws.row_dimensions[row].height = 22
        row += 2

        # Reset date column widths back
        for i in range(len(dates)):
            ws.column_dimensions[get_column_letter(2 + i)].width = 8
        # Reset from client_col onwards
        for col, w in col_w.items():
            if col >= 2:
                ws.column_dimensions[get_column_letter(col)].width = w

    # ══════════════════════════════════════════════
    # SECTION D: LABOUR ALLOCATION DETAILS
    # ══════════════════════════════════════════════
    ALLOC_COLS = 12
    if all_roles and dates:
        _section_banner(ws, row, "D.  LABOUR ALLOCATION DETAILS (Cost Breakdown per Role)", col_span=ALLOC_COLS)
        row += 1

        # ── Header: Role | Qty | DS group | NS group | Total $ | Client $ ──
        alloc_hdrs = ["Role", "Qty",
                      "DS Shifts", "DS hrs", "DS $/hr", "DS Cost",
                      "NS Shifts", "NS hrs", "NS $/hr", "NS Cost",
                      "Total $", "Client $"]
        for i, lbl in enumerate(alloc_hdrs):
            hc = ws.cell(row=row, column=1 + i, value=lbl)
            hc.fill = _fill(NAVY)
            hc.font = _font(bold=True, colour=WHITE, size=8)
            hc.alignment = _align("center", "center", wrap=True)
            hc.border = _border()
        ws.row_dimensions[row].height = 24
        row += 1

        total_labour = 0.0
        total_client_alloc = 0.0

        def _r(v):  # rate or blank when zero
            return v if v else None

        for r in all_roles:
            c = _calc_role_cost(r, manning, shift_map, dates, multiplier)
            total_labour       += c["total"]
            total_client_alloc += c["client_charge"]

            row_vals = [
                (f"{r.get('role','')} ×{r.get('qty',1)}", "left",  None,  False),
                (r.get("qty", 1),    "center", "0",   False),
                (c["ds_shifts"],     "center", "0",   False),
                (c["ds_hrs"],        "right",  "0.0", False),
                (_r(c["ds_rate"]),   "right",  '$#,##0.00', False),
                (c["ds_cost"],       "right",  '$#,##0.00', False),
                (c["ns_shifts"],     "center", "0",   False),
                (c["ns_hrs"],        "right",  "0.0", False),
                (_r(c["ns_rate"]),   "right",  '$#,##0.00', False),
                (c["ns_cost"],       "right",  '$#,##0.00', False),
                (c["total"],         "right",  '$#,##0.00', False),
                (c["client_charge"], "right",  '$#,##0.00', False),
            ]

            for col_idx, (val, halign, fmt, is_ot) in enumerate(row_vals, start=1):
                cell = ws.cell(row=row, column=col_idx, value="" if val is None else val)
                cell.border = _border()
                cell.alignment = _align(halign)
                if fmt:
                    cell.number_format = fmt
                if col_idx == ALLOC_COLS:        # Client $
                    cell.font = _font(size=10, bold=True, colour=NAVY)
                elif col_idx in (6, 10, 11):     # DS Cost / NS Cost / Total $
                    cell.font = _font(size=10, bold=True)
                elif col_idx == 1:
                    cell.font = _font(size=10)
                    cell.fill = _fill(GREY_LITE)
                else:
                    cell.font = _font(size=9)

            row += 1

        # Totals row
        ws.row_dimensions[row].height = 22
        tr = ws.cell(row=row, column=1, value="TOTALS")
        tr.fill = _fill(TEAL_DARK)
        tr.font = _font(bold=True, colour=WHITE, size=10)
        tr.border = _border()
        _merge(ws, row, 1, row, ALLOC_COLS - 2)

        tc_total = ws.cell(row=row, column=ALLOC_COLS - 1, value=total_labour)
        tc_total.number_format = '$#,##0.00'
        tc_total.font = _font(bold=True, size=11, colour=WHITE)
        tc_total.fill = _fill(TEAL_DARK)
        tc_total.border = _border()
        tc_total.alignment = _align("right")

        tc_client = ws.cell(row=row, column=ALLOC_COLS, value=total_client_alloc)
        tc_client.number_format = '$#,##0.00'
        tc_client.font = _font(bold=True, size=11, colour="FFE699")
        tc_client.fill = _fill(TEAL_DARK)
        tc_client.border = _border()
        tc_client.alignment = _align("right")

        row += 2

    # ══════════════════════════════════════════════
    # SECTION E: PRICING SUMMARY
    # ══════════════════════════════════════════════
    _section_banner(ws, row, "E.  PRICING SUMMARY", col_span=27)
    row += 1

    # Aggregate labour cost split by shift type across all roles
    agg = {"ds": 0.0, "ns": 0.0,
           "base": 0.0, "client": 0.0, "ds_shifts": 0, "ns_shifts": 0, "qty": 0}
    for r in all_roles:
        cd = _calc_role_cost(r, manning, shift_map, dates, multiplier)
        agg["ds"]     += cd["ds_cost"]
        agg["ns"]     += cd["ns_cost"]
        agg["base"]   += cd["total"]
        agg["client"] += cd["client_charge"]
        agg["ds_shifts"] += cd["ds_shifts"]
        agg["ns_shifts"] += cd["ns_shifts"]
        agg["qty"] += int(r.get("qty", 1) or 1)

    # Other costs — list of {label, amount}; tolerate legacy dict form
    if isinstance(oc, list):
        oc_items = [(str(o.get("label", "") or "Other Cost"), float(o.get("amount", 0) or 0)) for o in oc]
    elif isinstance(oc, dict):
        oc_items = [("Plant & Equipment", float(oc.get("plant", 0) or 0)),
                    ("Contractors",       float(oc.get("contractors", 0) or 0)),
                    ("Materials / Other", float(oc.get("materials", 0) or 0))]
    else:
        oc_items = []
    other_total = sum(a for _, a in oc_items)

    labour_client = agg["client"]
    markup        = agg["client"] - agg["base"]
    grand_total   = labour_client + other_total

    pricing_rows = [
        ("Labour Required (people)",   agg["qty"],            "0",         False, False),
        ("Total DS Shifts",            agg["ds_shifts"],      "0",         False, False),
        ("Total NS Shifts",            agg["ns_shifts"],      "0",         False, False),
        None,
        ("Day Shift Cost",             agg["ds"],             '$#,##0.00', False, False),
        ("Night Shift Cost",           agg["ns"],             '$#,##0.00', False, False),
        ("Total Labour (base)",        agg["base"],           '$#,##0.00', True,  False),
        ("Role Markup (Labour)",       markup,                '$#,##0.00', False, False),
        ("Labour Cost (Client)",       labour_client,         '$#,##0.00', True,  False),
        None,
    ]
    for lbl, amt in oc_items:
        pricing_rows.append((lbl, amt, '$#,##0.00', False, False))
    pricing_rows.append(("Total Other Costs", other_total, '$#,##0.00', True, False))
    pricing_rows.append(None)
    gst = grand_total * 0.10
    pricing_rows.append(("Project Total (ex GST)",  grand_total,       '$#,##0.00', True,  False))
    pricing_rows.append(("GST (10%)",               gst,               '$#,##0.00', False, False))
    pricing_rows.append(("PROJECT TOTAL (inc GST)", grand_total + gst, '$#,##0.00', True,  True))

    for item in pricing_rows:
        if item is None:
            row += 1
            continue
        label, value, fmt, is_subtotal, is_total = item
        ws.row_dimensions[row].height = 24 if is_total else 20

        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=3, value=value)

        if is_total:
            lc.fill = _fill(TEAL_DARK)
            lc.font = _font(bold=True, colour=WHITE, size=12)
            vc.fill = _fill(TEAL_DARK)
            vc.font = _font(bold=True, colour="FFE699", size=13)
        elif is_subtotal:
            lc.fill = _fill(GREY_BG)
            lc.font = _font(bold=True, size=10)
            vc.fill = _fill(GREY_BG)
            vc.font = _font(bold=True, size=11)
        else:
            lc.fill = _fill(GREY_LITE)
            lc.font = _font(size=10)
            vc.fill = _fill(WHITE)
            vc.font = _font(size=10)

        lc.border = _border()
        lc.alignment = _align("left", "center")
        _merge(ws, row, 1, row, 2)

        vc.number_format = fmt
        vc.border = _border()
        vc.alignment = _align("right", "center")
        _merge(ws, row, 3, row, 5)

        row += 1

    # ── Freeze panes ──────────────────────────────
    ws.freeze_panes = "A5"

    # ── Print settings ────────────────────────────
    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def make_file_name(job_id: str) -> str:
    date_str = datetime.now().strftime("%Y%m%d")
    return f"{job_id}_Quote_{date_str}.xlsx"
